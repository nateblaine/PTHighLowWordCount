from openpyxl import load_workbook

wb = load_workbook(filename='rawInput.xlsx')


# Models - move to dir later
class IntakeForm:
    def __init__(self, record_id, chief_complaint, group="high"):
        self.record_id = record_id
        self.chief_complaint = chief_complaint
        self.group = group.lower()

    def __str__(self):
        return ('Record ID: ' + str(self.record_id) + ' Chief Complaint: ' + str(self.chief_complaint))


class PresentationSymptom:
    def __init__(self, name, search_term, low_count=0, high_count=0, low_percent=0, high_percent=0):
        self.name = name
        self.search_term = search_term
        self.low_count = low_count
        self.high_count = high_count
        self.low_percent = low_percent
        self.high_percent = high_percent

    def __str__(self):
        return ('Presentation/Symptom: ' + str(self.name)
                + '\n' + ' Low Count: ' + str(self.low_count) + ' - ' + str(round(self.low_percent, 2)) + ' %'
                + '\n' + ' High Count: ' + str(self.high_count) + ' - ' + str(round(self.high_percent, 2)) + " %")


# Globals and Constants
low_sheet = "Low"
high_sheet = "High"

list_of_intake_forms = []

# Initial set up
list_of_symptom_counts = []
list_of_symptom_counts.append(PresentationSymptom('HTN',['htn','hypertension']))
list_of_symptom_counts.append(PresentationSymptom('Diabetes',['dm','diabetes']))
list_of_symptom_counts.append(PresentationSymptom('Gripe',['gripe']))
list_of_symptom_counts.append(PresentationSymptom('Pain',['pain']))
list_of_symptom_counts.append(PresentationSymptom('Shortness of breath',['shortness of breath','dyspnea', 'asthma', 'wheeze', 'cough', 'sob']))
list_of_symptom_counts.append(PresentationSymptom('Chest Pain',['chest pain']))
list_of_symptom_counts.append(PresentationSymptom('Headache',['headache']))
list_of_symptom_counts.append(PresentationSymptom('Rash',['itchy', 'itchiness', 'redness', 'rash']))
list_of_symptom_counts.append(PresentationSymptom('Blood Pressure Check',['bp', 'blood pressure']))
list_of_symptom_counts.append(PresentationSymptom('Dizziness',['dizziness', 'dizzy', 'lightheaded', 'vertigo']))
list_of_symptom_counts.append(PresentationSymptom('Medication Refill',['medication', 'meds', 'refill']))
list_of_symptom_counts.append(PresentationSymptom('Follow Up',['follow up']))
list_of_symptom_counts.append(PresentationSymptom('Fever',['fever']))
list_of_symptom_counts.append(PresentationSymptom('Diarrhea',['diarrhea']))
list_of_symptom_counts.append(PresentationSymptom('Constipation',['difficulty voiding', 'constipation']))
list_of_symptom_counts.append(PresentationSymptom('UTI',['urinary tract infection', 'uti', 'burning on urination']))
list_of_symptom_counts.append(PresentationSymptom('Loss of appetite',['loss of appetite']))
list_of_symptom_counts.append(PresentationSymptom('Weight loss',['weight loss']))
list_of_symptom_counts.append(PresentationSymptom('Insomnia',['difficulty sleeping', 'insomnia', 'not sleeping', 'trouble sleeping']))
list_of_symptom_counts.append(PresentationSymptom('Loss of vision',['loss of vision', 'blurry vision', 'vision issues']))
list_of_symptom_counts.append(PresentationSymptom('Stroke',['stroke']))
list_of_symptom_counts.append(PresentationSymptom('Numbness',['numbness', 'numb', 'tingling']))
list_of_symptom_counts.append(PresentationSymptom('Blood glucose',['blood glucose', 'blood sugar']))




# Helper funcs

def input_to_data():
    # Parse Data from raw input xls
    for row in ws.rows:
        temp_record_id = '';
        temp_chief_complaint = '';
        temp_group = ''

        for index, header in enumerate(headers):
            if header == 'Record ID':
                temp_record_id = row[index].value
            if header == 'Chief complaint':
                temp_chief_complaint = row[index].value
            if header == 'Group':
                temp_group = row[index].value

        if temp_record_id != None and temp_chief_complaint != None:
            list_of_intake_forms.append(IntakeForm(temp_record_id, temp_chief_complaint.lower(), temp_group if temp_group is not None else "high"))


def do_sheet_counts():
    # Loop through symptoms and search terms and make counter objects for final xls
    for symptom in list_of_symptom_counts:
        total_occurances = 0
        temp_low_count = 0;
        temp_high_count = 0;
        list_of_banned_rows = []

        for search_term in symptom.search_term:
            for intake_form in list_of_intake_forms:
                if search_term in intake_form.chief_complaint and intake_form.record_id not in list_of_banned_rows:
                    total_occurances += 1
                    if intake_form.group == "low":
                        temp_low_count += 1
                    elif intake_form.group == "high":
                        temp_high_count += 1;
                    list_of_banned_rows.append(intake_form.record_id)

        symptom.low_count += temp_low_count;
        symptom.low_percent = (symptom.low_count / total_occurances) * 100
        symptom.high_count += temp_high_count;
        symptom.high_percent = (symptom.high_count / total_occurances) * 100

# parse low sheet
ws = wb[low_sheet]
headers = [cell.value for cell in ws[1]]
input_to_data()
# parse high sheet
ws = wb[high_sheet]
headers = [cell.value for cell in ws[1]]
input_to_data()

print(list_of_intake_forms.__len__())

# count low sheet
do_sheet_counts()







# Print results to console
for x in list_of_symptom_counts:
    print(x)
